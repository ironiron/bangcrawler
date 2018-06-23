import requests, bs4
import openpyxl
import re
from bs4 import BeautifulSoup
url_pre=('https://www.banggood.com/Wholesale-Smartphones-c-1567-0-1-1-44-0_page')
url_post=('.html')
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = 'database'
iteration=0
#################################################
for i in range (1, 10):
    print('iteracja nr. '+str(i))
    res = requests.get(url_pre+str(i)+url_post)
    try:
        res.raise_for_status()
    except Exception as exc:
        print('There was a problem: %s' % (exc))
    exampleSoup = bs4.BeautifulSoup(res.text,'html.parser')
    type(exampleSoup)
    hyper=[0]
    k=0
    elems = exampleSoup.select('div ul li span > a')
    for j in range (1, len(elems)):
        if elems[j-1].attrs['href']!=elems[j].attrs['href'] and j>12:
            if (('reviews') in str(elems[j].attrs['href']))==0:
                hyper.append(elems[j])
                k=k+1
                sheet.cell(row=k+iteration*44, column=1).value = k+iteration*44
                sheet.cell(row=k+iteration*44, column=2).value = hyper[k].attrs['href']
    ##########################################
    for m in range (1, 45):
        res = requests.get(hyper[m].attrs['href'])
        try:
            res.raise_for_status()
        except Exception as exc:
            print('There was a problem: %s' % (exc))
        exampleSoup = bs4.BeautifulSoup(res.text,'html.parser')
        type(exampleSoup)
            ###########################################################################
        for n in range (1, 4):     
            elems = exampleSoup.select('div meta')
        try:
            price =float(elems[5].attrs['content'])
        except ValueError:
            try:
                price =float(elems[6].attrs['content'])
            except ValueError:
                price =float(elems[7].attrs['content'])
        print("m="+str(m))
        sheet.cell(row=m+iteration*44, column=3).value = float(price)
    ##############################
        links = []
        for link in exampleSoup(text=re.compile(r'([0-9]+mah)',re.IGNORECASE)):
            links.append(link.replace(u'\xa0', u' ').strip())
            aa=re.compile(r'([0-9]+mah)',re.IGNORECASE)
            mo=aa.search(links[0])
            battery=int(re.sub('[^0-9]','',mo.group()))
        sheet.cell(row=m+iteration*44, column=4).value = int(battery)
###########################################################
        ind=res.text.index('Dimensions')
        aa=re.compile(r'cm|mm',re.IGNORECASE)
        mo=aa.search(res.text[ind:])
        if mo.group() == 'cm':
            cm=1
        else:
            cm=0
        aa=re.compile(r'([0-9]{2}[0-9.]+)',re.IGNORECASE)
        mo=aa.search(res.text[ind:])
        d1=mo.group()
        ind=ind+mo.end()
        mo=aa.search(res.text[ind:])
        d2=mo.group()
        ind=ind+mo.end()
        mo=aa.search(res.text[ind:])
        d3=mo.group()
        if cm==1:
            sheet.cell(row=m+iteration*44, column=8).value = str(d1)+'*'+str(d2)+'*'+str(d3)+'*'+'cm'
            sheet.cell(row=m+iteration*44, column=6).value =float(d1)*float(d2)*float(d3)
        else:
            sheet.cell(row=m+iteration*44, column=8).value = str(d1)+'*'+str(d2)+'*'+str(d3)+'*'+'mm'
            sheet.cell(row=m+iteration*44, column=6).value =float(d1)*float(d2)*float(d3)/1000
        sheet.cell(row=m+iteration*44, column=9).value =float(d1)
        sheet.cell(row=m+iteration*44, column=10).value =float(d2)
        sheet.cell(row=m+iteration*44, column=11).value =float(d3)
    #####################################################
        for link in exampleSoup(text=re.compile(r'Earphone Port')):
            links.append(link)
            sheet.cell(row=m+iteration*44, column=5).value = 0
            if ("Earphone Port" in link) ==True:
                sheet.cell(row=m+iteration*44, column=5).value = 1
                break
        wb.save('banggood_final2.xlsx')
#################################################
    #sheet.cell(row=i, column=1).value = i
    #sheet.cell(row=i, column=2).value = elems[0].getText()
    iteration=iteration+1
    wb.save('banggood_final2.xlsx')
    

wb.save('banggood_final2.xlsx')
    
